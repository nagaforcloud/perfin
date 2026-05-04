"""
Excel Report Generation Module.

Generates comprehensive Excel reports with multiple sheets:
- All Transactions
- Categorized Expenses
- Income vs Expense Summary
- Recurring Payments
- Large Transactions
- Monthly Summary
- Financial Health

Uses openpyxl for Excel file creation with formatting and styling.
"""

import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, Color
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule

from core.config import Config
from core.ledger import Ledger
from .financial_summary import FinancialAnalyzer, FinancialSummary
from .detect_recurring import RecurringPaymentDetector
from .detect_anomalies import AnomalyDetector

logger = logging.getLogger(__name__)


# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
TOTAL_FONT = Font(bold=True)
CURRENCY_FORMAT = '#,##0.00'
DATE_FORMAT = 'yyyy-mm-dd'

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
RIGHT_ALIGNMENT = Alignment(horizontal='right', vertical='center')
LEFT_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)


@dataclass
class ReportConfig:
    """Configuration for Excel report generation."""
    include_all_transactions: bool = True
    include_category_summary: bool = True
    include_monthly_summary: bool = True
    include_recurring_payments: bool = True
    include_large_transactions: bool = True
    include_anomalies: bool = True
    include_financial_health: bool = True


class ExcelReportGenerator:
    """
    Generates comprehensive Excel financial reports.
    
    Creates a multi-sheet workbook with:
    - Formatted tables
    - Summary statistics
    - Conditional formatting
    - Charts (optional)
    """
    
    def __init__(
        self,
        config: Optional[Config] = None,
        ledger: Optional[Ledger] = None
    ):
        """
        Initialize the report generator.
        
        Args:
            config: Configuration object
            ledger: Ledger instance
        """
        self.config = config or Config()
        self.ledger = ledger or Ledger(self.config)
        self.wb: Optional[Workbook] = None
    
    def generate_report(
        self,
        output_path: Optional[Path] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        report_config: Optional[ReportConfig] = None
    ) -> Path:
        """
        Generate a comprehensive Excel report.
        
        Args:
            output_path: Output file path (auto-generated if not provided)
            start_date: Start of report period
            end_date: End of report period
            report_config: Report configuration options
        
        Returns:
            Path to generated report
        """
        config = report_config or ReportConfig()
        
        # Generate output path if not provided
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = self.config.reports_dir / f"financial_report_{timestamp}.xlsx"
        else:
            output_path = Path(output_path)
        
        logger.info(f"Generating Excel report: {output_path}")
        
        # Create workbook
        self.wb = Workbook()
        
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        
        # Generate each sheet
        if config.include_all_transactions:
            self._create_transactions_sheet(start_date, end_date)
        
        if config.include_category_summary:
            self._create_category_summary_sheet(start_date, end_date)
        
        if config.include_monthly_summary:
            self._create_monthly_summary_sheet(start_date, end_date)
        
        if config.include_recurring_payments:
            self._create_recurring_payments_sheet()
        
        if config.include_large_transactions:
            self._create_large_transactions_sheet(start_date, end_date)
        
        if config.include_anomalies:
            self._create_anomalies_sheet(start_date, end_date)
        
        if config.include_financial_health:
            self._create_financial_health_sheet(start_date, end_date)
        
        # Create summary sheet (always first)
        self._create_summary_sheet(start_date, end_date)
        
        # Move summary to first position
        summary_sheet = self.wb['Summary']
        self.wb.remove(summary_sheet)
        self.wb.add_sheet(summary_sheet, 0)
        
        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)
        
        logger.info(f"Report saved to: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(
        self,
        start_date: Optional[str],
        end_date: Optional[str]
    ):
        """Create the summary dashboard sheet."""
        ws = self.wb.create_sheet('Summary')
        
        # Get financial summary
        analyzer = FinancialAnalyzer(self.config, self.ledger)
        summary = analyzer.generate_summary(start_date, end_date)
        health = analyzer.get_financial_health_score()
        
        # Title
        ws['A1'] = 'Financial Summary Report'
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:C1')
        
        # Period
        ws['A3'] = 'Report Period:'
        ws['B3'] = f"{summary.period_start} to {summary.period_end}"
        
        # Key Metrics
        metrics = [
            ('Total Income', summary.total_income),
            ('Total Expenses', summary.total_expenses),
            ('Net Savings', summary.net_savings),
            ('Savings Rate', f"{summary.savings_rate}%"),
            ('Total Transactions', summary.transaction_count),
        ]
        
        ws['A5'] = 'Key Metrics'
        ws['A5'].font = HEADER_FONT
        
        row = 6
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if isinstance(value, (int, float)):
                ws[f'B{row}'].number_format = CURRENCY_FORMAT
            row += 1
        
        # Financial Health Score
        health_row = 15
        ws[f'A{health_row}'] = 'Financial Health Score'
        ws[f'A{health_row}'].font = HEADER_FONT
        
        health_row += 1
        ws[f'A{health_row}'] = 'Overall Score'
        ws[f'B{health_row}'] = f"{health['total_score']}/{health['max_score']}"
        ws[f'C{health_row}'] = health['rating']
        
        # Category Breakdown (Top 5)
        cat_row = 20
        ws[f'A{cat_row}'] = 'Top Expense Categories'
        ws[f'A{cat_row}'].font = HEADER_FONT
        
        # Sort categories by expense amount
        expense_categories = [
            (cat, data) for cat, data in summary.category_breakdown.items()
            if data['total'] < 0
        ]
        expense_categories.sort(key=lambda x: x[1]['total'])
        
        cat_row += 1
        for cat, data in expense_categories[:5]:
            ws[f'A{cat_row}'] = cat
            ws[f'B{cat_row}'] = abs(data['total'])
            ws[f'B{cat_row}'].number_format = CURRENCY_FORMAT
            ws[f'C{cat_row}'] = f"{data['percentage']}%"
            cat_row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
    
    def _create_transactions_sheet(
        self,
        start_date: Optional[str],
        end_date: Optional[str]
    ):
        """Create the all transactions sheet."""
        ws = self.wb.create_sheet('All Transactions')
        
        # Get transactions
        transactions = self.ledger.get_all_transactions(
            start_date=start_date,
            end_date=end_date
        )
        
        # Headers
        headers = ['ID', 'Date', 'Description', 'Category', 'Amount', 'Account']
        self._write_headers(ws, headers)
        
        # Data rows
        for idx, txn in enumerate(transactions, 2):
            ws[f'A{idx}'] = txn.get('id', '')
            ws[f'B{idx}'] = txn.get('date', '')
            ws[f'C{idx}'] = txn.get('description', '')
            ws[f'D{idx}'] = txn.get('category', 'Needs Review')
            amount = txn.get('amount', 0)
            ws[f'E{idx}'] = amount
            ws[f'E{idx}'].number_format = CURRENCY_FORMAT
            # Color code amounts
            if amount < 0:
                ws[f'E{idx}'].font = Font(color='FF0000')  # Red for expenses
            else:
                ws[f'E{idx}'].font = Font(color='00AA00')  # Green for income
            ws[f'F{idx}'] = txn.get('account', '')
        
        # Apply formatting
        self._apply_table_formatting(ws, len(transactions) + 1)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def _create_category_summary_sheet(
        self,
        start_date: Optional[str],
        end_date: Optional[str]
    ):
        """Create the category summary sheet."""
        ws = self.wb.create_sheet('Category Summary')
        
        # Get category summary from ledger
        categories = self.ledger.get_category_summary(
            start_date=start_date,
            end_date=end_date
        )
        
        # Headers
        headers = ['Category', 'Count', 'Total Amount', 'Total Debits', 'Total Credits']
        self._write_headers(ws, headers)
        
        # Data rows
        for idx, cat in enumerate(categories, 2):
            ws[f'A{idx}'] = cat.get('category', 'Unknown')
            ws[f'B{idx}'] = cat.get('count', 0)
            ws[f'C{idx}'] = cat.get('total_amount', 0)
            ws[f'C{idx}'].number_format = CURRENCY_FORMAT
            ws[f'D{idx}'] = cat.get('total_debits', 0)
            ws[f'D{idx}'].number_format = CURRENCY_FORMAT
            ws[f'E{idx}'] = cat.get('total_credits', 0)
            ws[f'E{idx}'].number_format = CURRENCY_FORMAT
        
        # Apply formatting
        self._apply_table_formatting(ws, len(categories) + 1)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _create_monthly_summary_sheet(
        self,
        start_date: Optional[str],
        end_date: Optional[str]
    ):
        """Create the monthly summary sheet."""
        ws = self.wb.create_sheet('Monthly Summary')
        
        # Get monthly summary from ledger
        monthly = self.ledger.get_monthly_summary()
        
        # Headers
        headers = ['Month', 'Income', 'Expenses', 'Net', 'Transaction Count']
        self._write_headers(ws, headers)
        
        # Data rows
        for idx, month in enumerate(monthly, 2):
            ws[f'A{idx}'] = month.get('month', '')
            ws[f'B{idx}'] = month.get('income', 0)
            ws[f'B{idx}'].number_format = CURRENCY_FORMAT
            ws[f'C{idx}'] = month.get('expenses', 0)
            ws[f'C{idx}'].number_format = CURRENCY_FORMAT
            net = month.get('net', 0)
            ws[f'D{idx}'] = net
            ws[f'D{idx}'].number_format = CURRENCY_FORMAT
            # Color code net
            if net < 0:
                ws[f'D{idx}'].font = Font(color='FF0000', bold=True)
            else:
                ws[f'D{idx}'].font = Font(color='00AA00', bold=True)
            ws[f'E{idx}'] = 0  # Would need to calculate separately
        
        # Apply formatting
        self._apply_table_formatting(ws, len(monthly) + 1)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _create_recurring_payments_sheet(self):
        """Create the recurring payments sheet."""
        ws = self.wb.create_sheet('Recurring Payments')
        
        # Detect recurring payments
        detector = RecurringPaymentDetector(self.config, self.ledger)
        recurring = detector.detect_all()
        
        if not recurring:
            ws['A1'] = 'No recurring payments detected'
            return
        
        # Headers
        headers = ['Merchant', 'Amount', 'Frequency', 'Occurrences', 
                   'Total Paid', 'First Date', 'Last Date', 'Category', 'Confidence']
        self._write_headers(ws, headers)
        
        # Data rows
        for idx, payment in enumerate(recurring, 2):
            ws[f'A{idx}'] = payment.description
            ws[f'B{idx}'] = abs(payment.amount)
            ws[f'B{idx}'].number_format = CURRENCY_FORMAT
            ws[f'C{idx}'] = payment.frequency
            ws[f'D{idx}'] = payment.occurrences
            ws[f'E{idx}'] = abs(payment.total_amount)
            ws[f'E{idx}'].number_format = CURRENCY_FORMAT
            ws[f'F{idx}'] = payment.first_date
            ws[f'G{idx}'] = payment.last_date
            ws[f'H{idx}'] = payment.category
            ws[f'I{idx}'] = f"{payment.confidence * 100:.0f}%"
        
        # Apply formatting
        self._apply_table_formatting(ws, len(recurring) + 1)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
    
    def _create_large_transactions_sheet(
        self,
        start_date: Optional[str],
        end_date: Optional[str]
    ):
        """Create the large transactions sheet."""
        ws = self.wb.create_sheet('Large Transactions')
        
        # Get large transactions
        transactions = self.ledger.get_large_transactions(
            threshold=self.config.large_transaction_threshold,
            start_date=start_date,
            end_date=end_date
        )
        
        if not transactions:
            ws['A1'] = f'No transactions above ${self.config.large_transaction_threshold:,.2f}'
            return
        
        # Headers
        headers = ['ID', 'Date', 'Description', 'Category', 'Amount', 'Account']
        self._write_headers(ws, headers)
        
        # Data rows
        for idx, txn in enumerate(transactions, 2):
            ws[f'A{idx}'] = txn.get('id', '')
            ws[f'B{idx}'] = txn.get('date', '')
            ws[f'C{idx}'] = txn.get('description', '')
            ws[f'D{idx}'] = txn.get('category', 'Needs Review')
            amount = txn.get('amount', 0)
            ws[f'E{idx}'] = amount
            ws[f'E{idx}'].number_format = CURRENCY_FORMAT
            ws[f'E{idx}'].font = Font(color='FF0000', bold=True)
            ws[f'F{idx}'] = txn.get('account', '')
        
        # Apply formatting
        self._apply_table_formatting(ws, len(transactions) + 1)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def _create_anomalies_sheet(
        self,
        start_date: Optional[str],
        end_date: Optional[str]
    ):
        """Create the anomalies sheet."""
        ws = self.wb.create_sheet('Anomalies')
        
        # Detect anomalies
        detector = AnomalyDetector(self.config, self.ledger)
        anomalies = detector.detect_all()
        
        if not anomalies:
            ws['A1'] = 'No anomalies detected'
            return
        
        # Headers
        headers = ['ID', 'Date', 'Description', 'Amount', 'Type', 'Severity', 'Reason']
        self._write_headers(ws, headers)
        
        # Data rows
        for idx, anomaly in enumerate(anomalies, 2):
            txn = anomaly.transaction
            ws[f'A{idx}'] = anomaly.transaction_id
            ws[f'B{idx}'] = txn.get('date', '')
            ws[f'C{idx}'] = txn.get('description', '')
            ws[f'D{idx}'] = txn.get('amount', 0)
            ws[f'D{idx}'].number_format = CURRENCY_FORMAT
            ws[f'E{idx}'] = anomaly.anomaly_type
            ws[f'F{idx}'] = anomaly.severity
            # Color code severity
            if anomaly.severity == 'critical':
                ws[f'F{idx}'].font = Font(color='FF0000', bold=True)
            elif anomaly.severity == 'high':
                ws[f'F{idx}'].font = Font(color='FF6600', bold=True)
            ws[f'G{idx}'] = anomaly.reason
        
        # Apply formatting
        self._apply_table_formatting(ws, len(anomalies) + 1)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50
    
    def _create_financial_health_sheet(
        self,
        start_date: Optional[str],
        end_date: Optional[str]
    ):
        """Create the financial health sheet."""
        ws = self.wb.create_sheet('Financial Health')
        
        analyzer = FinancialAnalyzer(self.config, self.ledger)
        health = analyzer.get_financial_health_score()
        
        # Title
        ws['A1'] = 'Financial Health Assessment'
        ws['A1'].font = TITLE_FONT
        
        # Overall Score
        ws['A3'] = 'Overall Score'
        ws['B3'] = f"{health['total_score']}/{health['max_score']}"
        ws['B3'].font = Font(bold=True, size=14)
        ws['C3'] = health['rating']
        
        # Metrics
        metrics = health.get('metrics', {})
        
        ws['A5'] = 'Component Scores'
        ws['A5'].font = HEADER_FONT
        
        metric_labels = {
            'savings_score': 'Savings Rate',
            'consistency_score': 'Expense Consistency',
            'budget_score': 'Budget Management',
            'categorization_score': 'Transaction Categorization'
        }
        
        row = 6
        for key, label in metric_labels.items():
            ws[f'A{row}'] = label
            ws[f'B{row}'] = f"{metrics.get(key, 0)}/25"
            row += 1
        
        # Recommendations
        rec_row = 15
        ws[f'A{rec_row}'] = 'Recommendations'
        ws[f'A{rec_row}'].font = HEADER_FONT
        
        recommendations = self._generate_recommendations(health, metrics)
        rec_row += 1
        for i, rec in enumerate(recommendations, rec_row):
            ws[f'A{i}'] = f"• {rec}"
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _generate_recommendations(
        self,
        health: Dict[str, Any],
        metrics: Dict[str, int]
    ) -> List[str]:
        """Generate financial health recommendations."""
        recommendations = []
        
        if metrics.get('savings_score', 0) < 15:
            recommendations.append(
                "Consider increasing your savings rate. Aim for at least 10-20% of income."
            )
        
        if metrics.get('consistency_score', 0) < 15:
            recommendations.append(
                "Your monthly expenses vary significantly. Consider creating a budget."
            )
        
        if metrics.get('budget_score', 0) < 15:
            recommendations.append(
                "Your expenses are close to or exceeding income. Review spending habits."
            )
        
        if metrics.get('categorization_score', 0) < 15:
            recommendations.append(
                "Many transactions are uncategorized. Review and categorize for better insights."
            )
        
        if not recommendations:
            recommendations.append(
                "Your financial health looks good! Continue monitoring your finances regularly."
            )
        
        return recommendations
    
    def _write_headers(self, ws, headers: List[str]):
        """Write header row with formatting."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER
    
    def _apply_table_formatting(self, ws, last_row: int):
        """Apply consistent table formatting."""
        # Apply borders to all cells
        for row in range(1, last_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
        
        # Alternate row colors
        light_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        for row in range(2, last_row + 1):
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = light_fill
        
        # Header alignment
        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).alignment = CENTER_ALIGNMENT


def generate_excel_report(
    output_path: Optional[Path] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    config: Optional[Config] = None,
    ledger: Optional[Ledger] = None
) -> Path:
    """
    Convenience function to generate Excel report.
    
    Args:
        output_path: Output file path
        start_date: Start of report period
        end_date: End of report period
        config: Configuration object
        ledger: Ledger instance
    
    Returns:
        Path to generated report
    """
    generator = ExcelReportGenerator(config=config, ledger=ledger)
    return generator.generate_report(output_path, start_date, end_date)
